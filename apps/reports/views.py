"""
apps/reports/views.py

Ten Excel (xlsx) report views — all staff-only.

Overall reports:
    1.  monthly_collection      — due / collected / outstanding by month
    2.  overdue_aging           — overdue installments bucketed by age
    3.  payment_plan_breakdown  — bookings & value by payment plan  (fixed double-count)
    4.  cash_flow               — NEW: monthly net cash (collected − expenses)
    5.  token_pipeline          — NEW: TOKEN bookings awaiting activation

Per-project reports (require project pk in URL):
    6.  project_plot_inventory  — plot status breakdown
    7.  project_revenue         — bookings revenue summary
    8.  project_expenses        — expenses by category
    9.  project_per_plot_detail — full per-plot customer & payment row

Per-customer reports:
    10. customer_statement      — NEW: full payment history for one customer

Bugs fixed vs previous version:
    • payment_plan_breakdown: two-query merge eliminates ORM fanout /
      double-counting that occurred when annotating across installments FK.
    • project_per_plot_detail: explicit is_deleted=False on bookings reverse
      relation (related managers bypass custom manager by default).
    • project_plot_inventory: SizeUnit choices pulled from class, not _meta.
    • All views: Booking.objects / Installment.objects already exclude
      soft-deleted rows via SoftDeleteManager — removed redundant
      .filter(is_deleted=False) calls.
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Count, Q, Sum, F
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone

from apps.bookings.models import Booking, Installment
from apps.customers.models import Customer
from apps.expenses.models import Expense
from apps.projects_and_plots.models import Plot, Project


# ─────────────────────────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────────────────────────

class _XL:
    """Minimal openpyxl wrapper used by every report."""

    DARK   = "1E293B"
    WHITE  = "FFFFFF"
    TOTAL  = "E2E8F0"
    ALT    = "F8FAFC"
    VIOLET = "7C3AED"
    GREEN  = "16A34A"
    AMBER  = "D97706"
    ROSE   = "E11D48"
    SKY    = "0284C7"
    SLATE  = "64748B"
    ORANGE = "EA580C"

    PKR_FMT  = '#,##0'
    DATE_FMT = 'DD-MMM-YYYY'

    def __init__(self, sheet_title: str = "Report"):
        self.wb       = openpyxl.Workbook()
        self.ws       = self.wb.active
        self.ws.title = sheet_title[:31]
        self.row      = 1

    def title(self, text: str, subtitle: str = "") -> None:
        c = self.ws.cell(self.row, 1, text)
        c.font = Font(bold=True, size=15, color=self.DARK)
        self.row += 1
        if subtitle:
            c2 = self.ws.cell(self.row, 1, subtitle)
            c2.font = Font(italic=True, size=10, color=self.SLATE)
            self.row += 1
        self.row += 1

    def section(self, text: str, n_cols: int = 1) -> None:
        c = self.ws.cell(self.row, 1, text)
        c.font      = Font(bold=True, size=11, color=self.WHITE)
        c.fill      = PatternFill("solid", fgColor=self.VIOLET)
        c.alignment = Alignment(vertical="center", indent=1)
        if n_cols > 1:
            self.ws.merge_cells(
                start_row=self.row, start_column=1,
                end_row=self.row,   end_column=n_cols,
            )
        self.ws.row_dimensions[self.row].height = 22
        self.row += 1

    def headers(self, cols: list[tuple[str, int]]) -> None:
        for i, (label, width) in enumerate(cols, 1):
            self.ws.column_dimensions[get_column_letter(i)].width = width
            c = self.ws.cell(self.row, i, label)
            c.font      = Font(bold=True, color=self.WHITE, size=10)
            c.fill      = PatternFill("solid", fgColor=self.DARK)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.ws.row_dimensions[self.row].height = 28
        self.row += 1

    def row_data(
        self,
        values: list,
        alt:   bool = False,
        bold:  bool = False,
        color: str | None = None,
    ) -> None:
        bg = color or (self.ALT if alt else self.WHITE)
        for i, val in enumerate(values, 1):
            c = self.ws.cell(self.row, i, val)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(vertical="center")
            if bold:
                c.font = Font(bold=True)
            if isinstance(val, (int, float)) and val > 999:
                c.number_format = self.PKR_FMT
            if isinstance(val, date):
                c.number_format = self.DATE_FMT
        self.row += 1

    def totals(self, values: list) -> None:
        self.row_data(values, bold=True, color=self.TOTAL)

    def blank(self) -> None:
        self.row += 1

    def response(self, filename: str) -> HttpResponse:
        buf = BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp


# ─────────────────────────────────────────────────────────────────
# UTILITIES
# ─────────────────────────────────────────────────────────────────

def _today_str() -> str:
    return timezone.localdate().strftime("%Y-%m-%d")


def _safe(val):
    return val or 0


def _pct(numerator, denominator) -> str:
    if not denominator:
        return "0%"
    return f"{numerator / denominator * 100:.1f}%"


# ─────────────────────────────────────────────────────────────────
# 1. MONTHLY COLLECTION REPORT
# ─────────────────────────────────────────────────────────────────

@staff_member_required
def monthly_collection(request):
    """
    One row per calendar month (all time, ascending).
    Columns: Month | # Installments | # Paid | # Overdue | # Pending | # Waived
             | Total Due | Total Collected | Outstanding | Collection %
    """
    rows = (
        Installment.objects
        .annotate(month=TruncMonth("due_date"))
        .values("month")
        .annotate(
            n_total   = Count("id"),
            n_paid    = Count("id", filter=Q(status=Installment.Status.PAID)),
            n_overdue = Count("id", filter=Q(status=Installment.Status.OVERDUE)),
            n_pending = Count("id", filter=Q(status=Installment.Status.PENDING)),
            n_waived  = Count("id", filter=Q(status=Installment.Status.WAIVED)),
            total_due = Sum("amount_due"),
            total_paid= Sum("amount_paid"),
        )
        .order_by("month")
    )

    xl = _XL("Monthly Collection")
    xl.title(
        "Monthly Collection Report",
        f"Generated: {_today_str()}  |  All projects",
    )
    xl.headers([
        ("Month",              16),
        ("# Installments",     14),
        ("# Paid",             10),
        ("# Overdue",          10),
        ("# Pending",          10),
        ("# Waived",           10),
        ("Total Due (₨)",      18),
        ("Collected (₨)",      18),
        ("Outstanding (₨)",    18),
        ("Collection %",       13),
    ])

    grand_due = grand_paid = 0

    for i, r in enumerate(rows):
        month_label = r["month"].strftime("%B %Y") if r["month"] else "—"
        due  = _safe(r["total_due"])
        paid = _safe(r["total_paid"])
        out  = due - paid
        grand_due  += due
        grand_paid += paid

        xl.row_data(
            [
                month_label,
                r["n_total"], r["n_paid"], r["n_overdue"],
                r["n_pending"], r["n_waived"],
                due, paid, out,
                _pct(paid, due),
            ],
            alt=bool(i % 2),
        )

    xl.blank()
    xl.totals([
        "TOTAL", "", "", "", "", "",
        grand_due, grand_paid,
        grand_due - grand_paid,
        _pct(grand_paid, grand_due),
    ])

    return xl.response(f"monthly_collection_{_today_str()}.xlsx")


# ─────────────────────────────────────────────────────────────────
# 2. OVERDUE AGING REPORT
# ─────────────────────────────────────────────────────────────────

@staff_member_required
def overdue_aging(request):
    """
    Every overdue installment with days-overdue and age bucket.
    Bucket summary table at the bottom.
    """
    today = timezone.localdate()

    overdue_qs = (
        Installment.objects
        .filter(status=Installment.Status.OVERDUE)
        .select_related("booking__customer", "booking__plot__project")
        .order_by("due_date")
    )

    def bucket(days: int) -> str:
        if days <= 30:  return "0–30 days"
        if days <= 60:  return "31–60 days"
        if days <= 90:  return "61–90 days"
        return "90+ days"

    bucket_order  = ["0–30 days", "31–60 days", "61–90 days", "90+ days"]
    bucket_totals = {b: {"count": 0, "due": 0, "paid": 0} for b in bucket_order}

    xl = _XL("Overdue Aging")
    xl.title(
        "Overdue Installment Aging Report",
        f"Generated: {_today_str()}  |  Reference date: {today}",
    )
    xl.headers([
        ("Customer",         22),
        ("CNIC",             18),
        ("Project",          20),
        ("Plot #",           10),
        ("Challan",          18),
        ("Due Date",         14),
        ("Days Overdue",     13),
        ("Age Bucket",       13),
        ("Amount Due (₨)",   18),
        ("Amount Paid (₨)",  18),
        ("Balance (₨)",      18),
    ])

    for i, inst in enumerate(overdue_qs):
        days    = (today - inst.due_date).days
        bkt     = bucket(days)
        balance = inst.amount_due - inst.amount_paid
        cust    = inst.booking.customer
        plot    = inst.booking.plot

        bucket_totals[bkt]["count"] += 1
        bucket_totals[bkt]["due"]   += inst.amount_due
        bucket_totals[bkt]["paid"]  += inst.amount_paid

        xl.row_data(
            [
                cust.full_name, cust.cnic,
                plot.project.name, plot.plot_number,
                inst.challan_number, inst.due_date,
                days, bkt,
                inst.amount_due, inst.amount_paid, balance,
            ],
            alt=bool(i % 2),
        )

    xl.blank()
    xl.blank()
    xl.section("Aging Summary", n_cols=5)
    xl.headers([
        ("Age Bucket",       16),
        ("# Installments",   15),
        ("Total Due (₨)",    18),
        ("Total Paid (₨)",   18),
        ("Balance (₨)",      18),
    ])
    grand = {"count": 0, "due": 0, "paid": 0}
    for bkt in bucket_order:
        t = bucket_totals[bkt]
        xl.row_data([bkt, t["count"], t["due"], t["paid"], t["due"] - t["paid"]])
        grand["count"] += t["count"]
        grand["due"]   += t["due"]
        grand["paid"]  += t["paid"]

    xl.totals([
        "TOTAL", grand["count"],
        grand["due"], grand["paid"],
        grand["due"] - grand["paid"],
    ])

    return xl.response(f"overdue_aging_{_today_str()}.xlsx")


# ─────────────────────────────────────────────────────────────────
# 3. PAYMENT PLAN BREAKDOWN  (fixed: two-query merge, no fanout)
# ─────────────────────────────────────────────────────────────────

@staff_member_required
def payment_plan_breakdown(request):
    """
    One row per payment plan with booking counts and financial totals.

    FIX: Previously used Sum("installments__amount_due") in a single
    annotated queryset which caused ORM fanout — aggregated booking-level
    values (total_price, token_amount, down_payment) were multiplied by
    the installment count. Now uses two separate queries merged by plan.
    """
    plan_labels = dict(Booking.PaymentPlan.choices)

    # ── Query 1: booking-level aggregates (no FK joins) ───────────
    booking_stats = {
        r["payment_plan"]: r
        for r in (
            Booking.objects
            .values("payment_plan")
            .annotate(
                n_bookings  = Count("id"),
                n_active    = Count("id", filter=Q(status=Booking.Status.ACTIVE)),
                n_completed = Count("id", filter=Q(status=Booking.Status.COMPLETED)),
                n_cancelled = Count("id", filter=Q(status=Booking.Status.CANCELLED)),
                n_token     = Count("id", filter=Q(status=Booking.Status.TOKEN)),
                total_value = Sum("total_price"),
                total_token = Sum("token_amount"),
                total_dp    = Sum("down_payment"),
            )
        )
    }

    # ── Query 2: installment-level aggregates joined to booking plan ─
    inst_stats = {
        r["booking__payment_plan"]: r
        for r in (
            Installment.objects
            .values("booking__payment_plan")
            .annotate(
                inst_due  = Sum("amount_due"),
                inst_paid = Sum("amount_paid"),
            )
        )
    }

    xl = _XL("Plan Breakdown")
    xl.title(
        "Payment Plan Breakdown Report",
        f"Generated: {_today_str()}",
    )
    xl.headers([
        ("Payment Plan",       18),
        ("# Bookings",         10),
        ("# Token",            10),
        ("# Active",           10),
        ("# Completed",        12),
        ("# Cancelled",        12),
        ("Total Value (₨)",    18),
        ("Token Total (₨)",    18),
        ("Down Payment (₨)",   18),
        ("Inst. Due (₨)",      18),
        ("Inst. Paid (₨)",     18),
        ("Outstanding (₨)",    18),
        ("Collection %",       13),
    ])

    g = {k: 0 for k in ["nb","nt","na","nc","ncn","tv","tt","dp","id_","ip"]}

    for i, plan in enumerate(Booking.PaymentPlan.values):
        b   = booking_stats.get(plan, {})
        ins = inst_stats.get(plan, {})
        due  = _safe(ins.get("inst_due"))
        paid = _safe(ins.get("inst_paid"))

        xl.row_data([
            plan_labels.get(plan, plan),
            _safe(b.get("n_bookings")),
            _safe(b.get("n_token")),
            _safe(b.get("n_active")),
            _safe(b.get("n_completed")),
            _safe(b.get("n_cancelled")),
            _safe(b.get("total_value")),
            _safe(b.get("total_token")),
            _safe(b.get("total_dp")),
            due, paid, due - paid,
            _pct(paid, due),
        ], alt=bool(i % 2))

        g["nb"]  += _safe(b.get("n_bookings"))
        g["nt"]  += _safe(b.get("n_token"))
        g["na"]  += _safe(b.get("n_active"))
        g["nc"]  += _safe(b.get("n_completed"))
        g["ncn"] += _safe(b.get("n_cancelled"))
        g["tv"]  += _safe(b.get("total_value"))
        g["tt"]  += _safe(b.get("total_token"))
        g["dp"]  += _safe(b.get("total_dp"))
        g["id_"] += due
        g["ip"]  += paid

    xl.blank()
    xl.totals([
        "TOTAL",
        g["nb"], g["nt"], g["na"], g["nc"], g["ncn"],
        g["tv"], g["tt"], g["dp"],
        g["id_"], g["ip"], g["id_"] - g["ip"],
        _pct(g["ip"], g["id_"]),
    ])

    return xl.response(f"payment_plan_breakdown_{_today_str()}.xlsx")


# ─────────────────────────────────────────────────────────────────
# 4. CASH FLOW REPORT  (NEW)
# ─────────────────────────────────────────────────────────────────

@staff_member_required
def cash_flow(request):
    """
    Monthly cash flow — money in vs money out vs net position.

    Money IN  = token payments received + down payments received
                + installments collected that month
    Money OUT = expenses recorded that month
    Net       = IN − OUT, with running balance
    """
    today  = timezone.localdate()

    # ── Installments collected by month ────────────────────────────
    inst_by_month = {
        r["month"]: _safe(r["total"])
        for r in (
            Installment.objects
            .filter(status=Installment.Status.PAID, paid_on__isnull=False)
            .annotate(month=TruncMonth("paid_on"))
            .values("month")
            .annotate(total=Sum("amount_paid"))
        )
    }

    # ── Token payments by month (token_received_on) ────────────────
    token_by_month = {
        r["month"]: _safe(r["total"])
        for r in (
            Booking.objects
            .filter(token_received_on__isnull=False, token_amount__gt=0)
            .annotate(month=TruncMonth("token_received_on"))
            .values("month")
            .annotate(total=Sum("token_amount"))
        )
    }

    # ── Down payments by month (down_payment_received_on) ─────────
    dp_by_month = {
        r["month"]: _safe(r["total"])
        for r in (
            Booking.objects
            .filter(down_payment_received_on__isnull=False, down_payment__gt=0)
            .annotate(month=TruncMonth("down_payment_received_on"))
            .values("month")
            .annotate(total=Sum("down_payment"))
        )
    }

    # ── Expenses by month ─────────────────────────────────────────
    exp_by_month = {
        r["month"]: _safe(r["total"])
        for r in (
            Expense.objects
            .annotate(month=TruncMonth("date"))
            .values("month")
            .annotate(total=Sum("amount"))
        )
    }

    # ── Merge all months into a sorted set ────────────────────────
    all_months = sorted(
        inst_by_month.keys() | token_by_month.keys()
        | dp_by_month.keys() | exp_by_month.keys()
    )

    xl = _XL("Cash Flow")
    xl.title(
        "Monthly Cash Flow Report",
        f"Generated: {_today_str()}  |  All projects",
    )
    xl.headers([
        ("Month",                16),
        ("Token Received (₨)",   18),
        ("Down Payments (₨)",    18),
        ("Installments (₨)",     18),
        ("Total In (₨)",         18),
        ("Expenses (₨)",         18),
        ("Net Cash (₨)",         18),
        ("Running Balance (₨)",  18),
    ])

    running = 0
    g_token = g_dp = g_inst = g_exp = 0

    for i, month in enumerate(all_months):
        token  = token_by_month.get(month, 0)
        dp     = dp_by_month.get(month, 0)
        inst   = inst_by_month.get(month, 0)
        exp    = exp_by_month.get(month, 0)
        total_in = token + dp + inst
        net      = total_in - exp
        running += net

        g_token += token
        g_dp    += dp
        g_inst  += inst
        g_exp   += exp

        xl.row_data(
            [
                month.strftime("%B %Y"),
                token, dp, inst, total_in,
                exp, net, running,
            ],
            alt=bool(i % 2),
        )

    g_in = g_token + g_dp + g_inst
    xl.blank()
    xl.totals([
        "TOTAL",
        g_token, g_dp, g_inst, g_in,
        g_exp, g_in - g_exp, "",   # running balance total doesn't sum
    ])

    return xl.response(f"cash_flow_{_today_str()}.xlsx")


# ─────────────────────────────────────────────────────────────────
# 5. TOKEN PIPELINE REPORT  (NEW)
# ─────────────────────────────────────────────────────────────────

@staff_member_required
def token_pipeline(request):
    """
    All TOKEN-status bookings awaiting down payment / activation.
    Shows how long each has been in token stage and how much is
    still needed to move to ACTIVE.
    """
    today = timezone.localdate()

    bookings = (
        Booking.objects
        .filter(status=Booking.Status.TOKEN)
        .select_related("customer", "plot__project", "booked_by")
        .order_by("token_received_on")
    )

    plan_labels = dict(Booking.PaymentPlan.choices)

    xl = _XL("Token Pipeline")
    xl.title(
        "Token Pipeline — Pending Activation",
        f"Generated: {_today_str()}  |  Bookings awaiting down payment",
    )
    xl.headers([
        ("Customer",             22),
        ("CNIC",                 18),
        ("Phone",                15),
        ("Project",              20),
        ("Plot #",               10),
        ("Plot Price (₨)",       18),
        ("Payment Plan",         14),
        ("Token Date",           14),
        ("Days in Token",        13),
        ("Token Amount (₨)",     18),
        ("Down Payment Due (₨)", 20),
        ("Total Upfront Due (₨)",20),
        ("Booked By",            18),
    ])

    total_token_amt = 0
    total_dp_due    = 0

    for i, b in enumerate(bookings):
        days_in_token = (today - b.token_received_on).days if b.token_received_on else "—"
        cust          = b.customer
        upfront_due   = b.down_payment  # amount still needed to activate

        total_token_amt += _safe(b.token_amount)
        total_dp_due    += _safe(b.down_payment)

        xl.row_data(
            [
                cust.full_name,
                cust.cnic,
                cust.phone,
                b.plot.project.name,
                b.plot.plot_number,
                b.plot.price,
                plan_labels.get(b.payment_plan, b.payment_plan),
                b.token_received_on,
                days_in_token,
                b.token_amount,
                upfront_due,
                b.token_amount + upfront_due,
                b.booked_by.get_full_name() if b.booked_by else "—",
            ],
            alt=bool(i % 2),
        )

    xl.blank()
    xl.totals([
        f"TOTAL ({bookings.count()} bookings)",
        "", "", "", "", "",
        "", "", "",
        total_token_amt,
        total_dp_due,
        total_token_amt + total_dp_due,
        "",
    ])

    return xl.response(f"token_pipeline_{_today_str()}.xlsx")


# ─────────────────────────────────────────────────────────────────
# 6. PROJECT — PLOT INVENTORY  (fixed: SizeUnit from class not _meta)
# ─────────────────────────────────────────────────────────────────

@staff_member_required
def project_plot_inventory(request, pk: int):
    project = get_object_or_404(Project, pk=pk)
    plots   = (
        Plot.objects
        .filter(project=project)
        .order_by("block", "plot_number")
    )

    status_counts = plots.values("status").annotate(n=Count("id"), value=Sum("price"))
    summary       = {r["status"]: r for r in status_counts}

    xl = _XL("Plot Inventory")
    xl.title(
        f"Plot Inventory — {project.name}",
        f"Generated: {_today_str()}  |  Code: {project.code}  |  Location: {project.location}",
    )

    xl.section("Inventory Summary", n_cols=4)
    xl.headers([
        ("Status",           14),
        ("# Plots",          10),
        ("Total Value (₨)",  20),
        ("% of Total",       12),
    ])
    total_plots   = plots.count()
    status_order  = [Plot.Status.AVAILABLE, Plot.Status.TOKEN, Plot.Status.BOOKED, Plot.Status.SOLD]
    status_labels = dict(Plot.Status.choices)

    for s in status_order:
        r   = summary.get(s, {})
        n   = r.get("n", 0)
        v   = _safe(r.get("value"))
        pct = _pct(n, total_plots)
        xl.row_data([status_labels.get(s, s), n, v, pct])

    xl.blank()
    xl.section("Plot Detail", n_cols=7)
    xl.headers([
        ("Plot #",        10),
        ("Block",          8),
        ("Size",          12),
        ("Category",      14),
        ("Status",        12),
        ("Price (₨)",     18),
        ("Notes",         30),
    ])

    # FIX: use Plot.SizeUnit.choices instead of _meta field introspection
    size_unit_labels = dict(Plot.SizeUnit.choices)
    cat_labels       = dict(Plot.Category.choices)

    for i, plot in enumerate(plots):
        size_label = size_unit_labels.get(plot.size_unit, plot.size_unit)
        xl.row_data([
            plot.plot_number,
            plot.block or "",
            f"{plot.size} {size_label}",
            cat_labels.get(plot.category, plot.category),
            status_labels.get(plot.status, plot.status),
            plot.price,
            plot.notes or "",
        ], alt=bool(i % 2))

    return xl.response(f"{project.code}_plot_inventory_{_today_str()}.xlsx")


# ─────────────────────────────────────────────────────────────────
# 7. PROJECT — REVENUE REPORT
# ─────────────────────────────────────────────────────────────────

@staff_member_required
def project_revenue(request, pk: int):
    project  = get_object_or_404(Project, pk=pk)
    bookings = (
        Booking.objects
        .filter(plot__project=project)
        .select_related("customer", "plot", "booked_by")
        .prefetch_related("installments")
        .order_by("-booking_date")
    )

    plan_labels = dict(Booking.PaymentPlan.choices)

    xl = _XL("Revenue")
    xl.title(
        f"Revenue Report — {project.name}",
        f"Generated: {_today_str()}  |  Code: {project.code}",
    )
    xl.headers([
        ("Booking #",          10),
        ("Customer",           22),
        ("CNIC",               18),
        ("Plot #",             10),
        ("Booking Date",       14),
        ("Status",             12),
        ("Payment Plan",       14),
        ("Total Price (₨)",    18),
        ("Token (₨)",          14),
        ("Down Payment (₨)",   16),
        ("Inst. Due (₨)",      16),
        ("Inst. Paid (₨)",     16),
        ("Outstanding (₨)",    16),
        ("Collection %",       13),
    ])

    g = {k: 0 for k in ["tp", "tk", "dp", "id_", "ip"]}

    for i, b in enumerate(bookings):
        insts     = list(b.installments.all())
        inst_due  = sum(x.amount_due  for x in insts)
        inst_paid = sum(x.amount_paid for x in insts)

        xl.row_data([
            b.pk,
            b.customer.full_name,
            b.customer.cnic,
            b.plot.plot_number,
            b.booking_date,
            b.get_status_display(),
            plan_labels.get(b.payment_plan, b.payment_plan),
            b.total_price,
            b.token_amount,
            _safe(b.down_payment),
            inst_due,
            inst_paid,
            inst_due - inst_paid,
            _pct(inst_paid, inst_due),
        ], alt=bool(i % 2))

        g["tp"]  += b.total_price
        g["tk"]  += b.token_amount
        g["dp"]  += _safe(b.down_payment)
        g["id_"] += inst_due
        g["ip"]  += inst_paid

    xl.blank()
    xl.totals([
        "", "", "", "", "", "", "TOTAL",
        g["tp"], g["tk"], g["dp"],
        g["id_"], g["ip"], g["id_"] - g["ip"],
        _pct(g["ip"], g["id_"]),
    ])

    return xl.response(f"{project.code}_revenue_{_today_str()}.xlsx")


# ─────────────────────────────────────────────────────────────────
# 8. PROJECT — EXPENSES REPORT
# ─────────────────────────────────────────────────────────────────

@staff_member_required
def project_expenses(request, pk: int):
    project  = get_object_or_404(Project, pk=pk)
    expenses = (
        Expense.objects
        .filter(project=project)
        .select_related("submitted_by")
        .order_by("date")
    )

    cat_summary = (
        expenses.values("category")
        .annotate(n=Count("id"), total=Sum("amount"))
        .order_by("-total")
    )
    cat_labels = dict(Expense.Category.choices)
    pm_labels  = dict(Expense.PaymentMethod.choices)

    xl = _XL("Expenses")
    xl.title(
        f"Expenses Report — {project.name}",
        f"Generated: {_today_str()}  |  Code: {project.code}",
    )

    xl.section("By Category", n_cols=3)
    xl.headers([
        ("Category",       20),
        ("# Expenses",     12),
        ("Total (₨)",      18),
    ])
    grand_total = 0
    for r in cat_summary:
        label = cat_labels.get(r["category"], r["category"])
        xl.row_data([label, r["n"], r["total"]])
        grand_total += _safe(r["total"])

    xl.totals(["TOTAL", expenses.count(), grand_total])
    xl.blank()

    xl.section("Expense Detail", n_cols=8)
    xl.headers([
        ("Date",             14),
        ("Category",         18),
        ("Vendor",           22),
        ("Description",      35),
        ("Amount (₨)",       18),
        ("Payment Method",   16),
        ("Reference #",      18),
        ("Submitted By",     20),
    ])

    for i, exp in enumerate(expenses):
        xl.row_data([
            exp.date,
            cat_labels.get(exp.category, exp.category),
            exp.vendor_name or "",
            exp.description or "",
            exp.amount,
            pm_labels.get(exp.payment_method, exp.payment_method),
            exp.reference_number or "",
            exp.submitted_by.get_full_name() if exp.submitted_by_id else "",
        ], alt=bool(i % 2))

    return xl.response(f"{project.code}_expenses_{_today_str()}.xlsx")


# ─────────────────────────────────────────────────────────────────
# 9. PROJECT — PER-PLOT DETAIL  (fixed: explicit is_deleted filter)
# ─────────────────────────────────────────────────────────────────

@staff_member_required
def project_per_plot_detail(request, pk: int):
    """
    One row per plot, whether booked or not.

    FIX: plot.bookings.all() goes through the related manager which
    does NOT automatically apply SoftDeleteManager's filter. Added
    explicit .filter(is_deleted=False) to exclude soft-deleted bookings.
    """
    project = get_object_or_404(Project, pk=pk)
    plots   = (
        Plot.objects
        .filter(project=project)
        .prefetch_related("bookings__customer", "bookings__installments")
        .order_by("block", "plot_number")
    )

    plan_labels      = dict(Booking.PaymentPlan.choices)
    size_unit_labels = dict(Plot.SizeUnit.choices)
    cat_labels       = dict(Plot.Category.choices)
    status_labels    = dict(Plot.Status.choices)

    xl = _XL("Per-Plot Detail")
    xl.title(
        f"Per-Plot Detail — {project.name}",
        f"Generated: {_today_str()}  |  Code: {project.code}",
    )
    xl.headers([
        ("Plot #",           10),
        ("Block",             8),
        ("Size",             12),
        ("Category",         14),
        ("Plot Status",      12),
        ("Price (₨)",        16),
        ("Customer",         22),
        ("CNIC",             18),
        ("Phone",            15),
        ("Payment Plan",     14),
        ("Booking Date",     14),
        ("Token (₨)",        14),
        ("Down Payment (₨)", 16),
        ("Inst. Due (₨)",    16),
        ("Inst. Paid (₨)",   16),
        ("Outstanding (₨)",  16),
        ("Collection %",     13),
        ("Booking Status",   14),
    ])

    g = {k: 0 for k in ["price", "tk", "dp", "id_", "ip"]}

    for i, plot in enumerate(plots):
        size_label = size_unit_labels.get(plot.size_unit, plot.size_unit)
        cat_label  = cat_labels.get(plot.category, plot.category)
        stat_label = status_labels.get(plot.status, plot.status)

        g["price"] += _safe(plot.price)

        # FIX: must filter is_deleted=False explicitly on reverse relation
        active_booking = next(
            (
                b for b in plot.bookings.all()
                if not b.is_deleted
                and b.status not in (Booking.Status.CANCELLED,)
            ),
            None,
        )

        if active_booking:
            b         = active_booking
            insts     = list(b.installments.filter(is_deleted=False))
            inst_due  = sum(x.amount_due  for x in insts)
            inst_paid = sum(x.amount_paid for x in insts)
            cust      = b.customer

            g["tk"]  += _safe(b.token_amount)
            g["dp"]  += _safe(b.down_payment)
            g["id_"] += inst_due
            g["ip"]  += inst_paid

            xl.row_data([
                plot.plot_number, plot.block or "",
                f"{plot.size} {size_label}", cat_label, stat_label,
                plot.price,
                cust.full_name, cust.cnic, cust.phone,
                plan_labels.get(b.payment_plan, b.payment_plan),
                b.booking_date,
                b.token_amount, _safe(b.down_payment),
                inst_due, inst_paid, inst_due - inst_paid,
                _pct(inst_paid, inst_due),
                b.get_status_display(),
            ], alt=bool(i % 2))

        else:
            xl.row_data([
                plot.plot_number, plot.block or "",
                f"{plot.size} {size_label}", cat_label, stat_label,
                plot.price,
                "— Vacant —", "", "",
                "", "",
                0, 0, 0, 0, 0, "0%", "—",
            ], alt=bool(i % 2))

    xl.blank()
    xl.totals([
        "", "", "", "", "TOTAL", g["price"],
        "", "", "", "", "",
        g["tk"], g["dp"],
        g["id_"], g["ip"], g["id_"] - g["ip"],
        _pct(g["ip"], g["id_"]),
        "",
    ])

    return xl.response(f"{project.code}_per_plot_detail_{_today_str()}.xlsx")


# ─────────────────────────────────────────────────────────────────
# 10. CUSTOMER STATEMENT  (NEW)
# ─────────────────────────────────────────────────────────────────

@staff_member_required
def customer_statement(request, pk: int):
    """
    Full payment history for one customer.
    Section per booking, installments listed within each section.
    Summary box at top: total booked value, total paid, outstanding.
    """
    customer = get_object_or_404(Customer, pk=pk)
    bookings = (
        Booking.objects
        .filter(customer=customer)
        .select_related("plot__project", "booked_by")
        .prefetch_related("installments")
        .order_by("booking_date")
    )

    plan_labels = dict(Booking.PaymentPlan.choices)
    inst_status_labels = dict(Installment.Status.choices)

    xl = _XL("Customer Statement")
    xl.title(
        f"Customer Statement — {customer.full_name}",
        (
            f"CNIC: {customer.cnic}  |  "
            f"Phone: {customer.phone}  |  "
            f"Generated: {_today_str()}"
        ),
    )

    # ── Customer summary box ──────────────────────────────────────
    all_bookings = list(bookings)
    grand_value  = sum(_safe(b.total_price) for b in all_bookings)
    grand_token  = sum(_safe(b.token_amount) for b in all_bookings)
    grand_dp     = sum(_safe(b.down_payment) for b in all_bookings)
    grand_inst_due = grand_inst_paid = 0

    for b in all_bookings:
        insts = list(b.installments.filter(is_deleted=False))
        grand_inst_due  += sum(x.amount_due  for x in insts)
        grand_inst_paid += sum(x.amount_paid for x in insts)

    total_paid       = grand_token + grand_dp + grand_inst_paid
    total_outstanding= grand_value - total_paid

    xl.section("Account Summary", n_cols=4)
    xl.headers([
        ("Metric",               25),
        ("Amount (₨)",           20),
        ("",                      1),
        ("",                      1),
    ])
    xl.row_data(["Total Bookings",        len(all_bookings), "", ""])
    xl.row_data(["Total Booked Value",    grand_value,       "", ""])
    xl.row_data(["Total Token Received",  grand_token,       "", ""])
    xl.row_data(["Total Down Payment",    grand_dp,          "", ""])
    xl.row_data(["Total Installments Paid", grand_inst_paid, "", ""])
    xl.row_data(["Total Paid (all heads)", total_paid,       "", ""], bold=True)
    xl.row_data(["Outstanding Balance",   total_outstanding, "", ""], bold=True)

    xl.blank()
    xl.blank()

    # ── One section per booking ───────────────────────────────────
    for b in all_bookings:
        insts     = list(b.installments.filter(is_deleted=False).order_by("installment_number"))
        inst_due  = sum(x.amount_due  for x in insts)
        inst_paid = sum(x.amount_paid for x in insts)
        upfront   = _safe(b.token_amount) + _safe(b.down_payment)

        xl.section(
            f"Booking #{b.pk}  |  Plot {b.plot.plot_number}  "
            f"|  {b.plot.project.name}  |  {b.get_status_display()}",
            n_cols=9,
        )

        # Booking summary row
        xl.headers([
            ("Booking Date",     14),
            ("Payment Plan",     14),
            ("Total Price (₨)",  18),
            ("Token (₨)",        14),
            ("Down Payment (₨)", 18),
            ("Inst. Due (₨)",    16),
            ("Inst. Paid (₨)",   16),
            ("Outstanding (₨)",  16),
            ("Collection %",     13),
        ])
        xl.row_data([
            b.booking_date,
            plan_labels.get(b.payment_plan, b.payment_plan),
            b.total_price,
            b.token_amount,
            _safe(b.down_payment),
            inst_due, inst_paid,
            inst_due - inst_paid,
            _pct(inst_paid, inst_due),
        ])

        xl.blank()

        # Installment rows
        if insts:
            xl.headers([
                ("#",             6),
                ("Challan",      18),
                ("Due Date",     14),
                ("Amount Due (₨)",18),
                ("Amount Paid (₨)",18),
                ("Balance (₨)",  16),
                ("Paid On",      14),
                ("Status",       12),
                ("Notes",        25),
            ])
            for j, inst in enumerate(insts):
                xl.row_data([
                    inst.installment_number,
                    inst.challan_number,
                    inst.due_date,
                    inst.amount_due,
                    inst.amount_paid,
                    inst.amount_due - inst.amount_paid,
                    inst.paid_on or "—",
                    inst_status_labels.get(inst.status, inst.status),
                    inst.notes or "",
                ], alt=bool(j % 2))
        else:
    
            xl.row_data(["No installments generated yet.", "", "", "", "", "", "", "", ""])

        xl.blank()
        xl.blank()

    return xl.response(f"customer_statement_{customer.cnic}_{_today_str()}.xlsx")